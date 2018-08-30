from openpyxl import load_workbook
import os
import sys
import DataValidation


class ExcelFunctions:

    def __init__(self):
        # self.directory = r'C:\Users\599247\Desktop\Scripts'
        self.directory = os.path.expanduser('~\Desktop\Scripts')
        self.file_name = 'SSH.xlsx'
        self.file_and_path = os.path.join(self.directory, self.file_name)
        self.wb, self.ws = self.open_workbook(self.directory, self.file_name)

        self.host_info = {}
        self.create_dict()
        self.current_ip = None

    @staticmethod
    def open_workbook(directory, file_name):
        if os.path.exists(directory):
            os.chdir(directory)  # Change working directory
        else:
            print(f"ERROR: Directory not found: {directory}\nBut I'll make it for you anyway :)")
            os.mkdir(directory)

        try:
            wb = load_workbook(filename=file_name)
            #ws = wb["Sheet1"]
            ws = wb.worksheets[0]
            return wb, ws
        except FileNotFoundError:
            print(f'ERROR: {file_name} not found in {directory}')
            sys.exit()

    def save_workbook(self, attempt=1):  # Saves a Excel Workbook using openpyxl
        wb = self.wb
        os.chdir(self.directory)
        name = self.file_name.split(".xlsx")[0]

        # os.system("TASKKILL /F /IM EXCEL.EXE")
        argument = '...'
        # os.system()
        # wb.save(filename=self.fileName)

        while attempt >= 1:
            try:
                if attempt == 1:
                    wb.save(filename=name + '.xlsx')
                    print(f'Workbook saved as {name}.xlsx')
                else:
                    wb.save(filename=name + f'({attempt}).xlsx')
                    print(f'Workbook saved as {name}({attempt}).xlsx')
            except PermissionError:
                attempt += 1
                print(f'Cannot overwrite workbook. Is it open? Attempting to save as {name} ({attempt}).xlsx')
                self.save_workbook(attempt)
            break

    def get_col_list_by_header(self, header_name):
        ws = self.ws
        items = []
        try:
            for colNum in range(1, ws.max_column + 1):
                if header_name in ws.cell(row=1, column=colNum).value:
                    for rowNum in range(2, ws.max_row + 1):
                        items.append(ws.cell(row=rowNum, column=colNum).value)
                    break
        except TypeError as err:
            pass
        return items

    def get_col_cell_by_header(self, header_name):
        ws = self.ws
        items = []
        try:
            for colNum in range(1, ws.max_column + 1):
                if header_name in ws.cell(row=1, column=colNum).value:
                    for rowNum in range(2, ws.max_row + 1):
                        items.append(ws.cell(row=rowNum, column=colNum))
                    break
        except TypeError as err:
            pass
        return items

    def get_cell(self, header_name, start_cell):
        ws = self.ws
        cell = None
        try:
            for colNum in range(1, ws.max_column + 1):
                if header_name in ws.cell(row=1, column=colNum).value:
                        cell = ws.cell(row=start_cell.row, column=colNum)
        except TypeError as err:
            pass
        if cell is None:
            print(f"'{header_name}' was not found as a column header in {self.file_name}")
        return cell

    def create_dict(self):
        list_of_cells_by_ip = self.get_col_cell_by_header("IP Address")
        ip_list = self.get_col_list_by_header("IP Address")
        vendor_list = self.get_col_list_by_header("Vendor")
        run_list = self.get_col_list_by_header("Run?")

        for i in range(0, len(ip_list)):
            if None not in {ip_list[i], run_list[i], vendor_list[i]}:
                if "N" not in run_list[i] and "n" not in run_list[i]:
                    ip = ip_list[i]
                    if DataValidation.is_valid_ipv4(ip) is True:  # Only runs if it is a 'real' IP
                        self.host_info[ip] = {}
                        self.host_info[ip]["ip"] = ip
                        self.host_info[ip]["vendor"] = vendor_list[i]
                        self.host_info[ip]["run"] = run_list[i]
                        self.host_info[ip]["cell"] = list_of_cells_by_ip[i]
        return self.host_info


if __name__ == '__main__':
    pass
    #print('Excel Functions is being run by itself')
    #ExcelFunctions = ExcelFunctions()
else:
    pass
    #print('Excel Functions is being imported from another module')
    #ExcelFunctions = ExcelFunctions()

